import os
import openpyxl
import psycopg2

from import_utils import parse_followers

DATABASE_URL = os.environ.get('DATABASE_URL')

def import_influencers(cursor, wb):
    influencer_sheets = {
        'Collabs': 'Collabs',
        'Lifestyle': 'Lifestyle', 
        'Fashion': 'Fashion',
        'Home': 'Home',
        'Beauty': 'Beauty',
        'Food ': 'Food',
        'Car Reviews': 'Car Reviews',
        'Moms': 'Moms',
    }
    
    inserted = 0
    for sheet_name, category in influencer_sheets.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found, skipping...")
            continue
            
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        print(f"Processing sheet: {sheet_name}, headers: {headers[:7]}")
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
                
            name = str(row[0]).strip() if row[0] else None
            if not name:
                continue
            
            tiktok_url = None
            instagram_url = None
            followers = None
            niche = None
            phone = None
            notes = None
            
            for i, header in enumerate(headers):
                if i >= len(row):
                    break
                val = row[i]
                if not val:
                    continue
                    
                header_lower = str(header).lower() if header else ''
                
                if 'tiktok' in header_lower or 'username' in header_lower:
                    if 'tiktok.com' in str(val):
                        tiktok_url = str(val)
                elif 'instagram' in header_lower:
                    if 'instagram.com' in str(val):
                        instagram_url = str(val)
                elif 'follower' in header_lower:
                    followers = parse_followers(val)
                elif 'industry' in header_lower or 'niche' in header_lower:
                    niche = str(val) if val else None
                elif 'phone' in header_lower or 'contact' in header_lower:
                    if val and str(val).replace('.', '').replace('-', '').isdigit():
                        phone = str(int(float(val))) if isinstance(val, float) else str(val)
                elif 'comment' in header_lower or 'rate' in header_lower:
                    notes = str(val) if val else None
            
            if not niche:
                niche = category
                
            cursor.execute("""
                INSERT INTO influencers (name, tiktok_url, instagram_url, followers, niche, phone, category, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (name, tiktok_url, instagram_url, followers, niche, phone, category, notes))
            inserted += 1
    
    return inserted

def import_ugc_creators(cursor, wb):
    if 'UGC' not in wb.sheetnames:
        print("UGC sheet not found")
        return 0
        
    ws = wb['UGC']
    headers = [cell.value for cell in ws[1]]
    print(f"UGC headers: {headers}")
    
    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
            
        name = str(row[0]).strip() if row[0] else None
        if not name:
            continue
        
        phone = None
        handle = None
        niche = None
        has_mock_video = False
        portfolio_url = None
        age = None
        gender = None
        languages = None
        accepts_gifted = False
        turnaround = None
        has_equipment = False
        has_editing = False
        can_voiceover = False
        skills_rating = None
        base_rate = None
        
        for i, header in enumerate(headers):
            if i >= len(row):
                break
            val = row[i]
            if val is None:
                continue
                
            header_lower = str(header).lower() if header else ''
            
            if 'number' in header_lower and 'follower' not in header_lower:
                if val and str(val).replace('.', '').replace('-', '').isdigit():
                    phone = str(int(float(val))) if isinstance(val, float) else str(val)
            elif 'handle' in header_lower:
                handle = str(val) if val else None
            elif 'niche' in header_lower:
                niche = str(val) if val else None
            elif 'mock' in header_lower:
                has_mock_video = bool(val)
            elif 'portfolio' in header_lower:
                portfolio_url = str(val) if val else None
            elif 'age' in header_lower:
                try:
                    age = int(float(val)) if val and str(val).replace('.','').isdigit() else None
                except (ValueError, TypeError):
                    age = None
            elif 'gender' in header_lower:
                gender = str(val) if val else None
            elif 'language' in header_lower:
                languages = str(val) if val else None
            elif 'gifted' in header_lower:
                accepts_gifted = bool(val)
            elif 'turnaround' in header_lower:
                turnaround = str(val) if val else None
            elif 'equipment' in header_lower:
                has_equipment = bool(val)
            elif 'editing' in header_lower:
                has_editing = bool(val)
            elif 'voiceover' in header_lower:
                can_voiceover = bool(val)
            elif 'rating' in header_lower:
                skills_rating = str(val) if val else None
            elif 'rate' in header_lower and 'rating' not in header_lower:
                base_rate = str(val) if val else None
        
        cursor.execute("""
            INSERT INTO ugc_creators (name, phone, handle, niche, has_mock_video, portfolio_url, age, gender, 
                                       languages, accepts_gifted_collab, turnaround_time, has_equipment, 
                                       has_editing_skills, can_voiceover, skills_rating, base_rate)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (name, phone, handle, niche, has_mock_video, portfolio_url, age, gender, 
              languages, accepts_gifted, turnaround, has_equipment, has_editing, can_voiceover, 
              skills_rating, base_rate))
        inserted += 1
    
    return inserted

def main():
    print("Loading Excel file...")
    wb = openpyxl.load_workbook('attached_assets/Kreate&co_Creator_Network_1770117705423.xlsx')
    print(f"Sheets: {wb.sheetnames}")
    
    print("\nConnecting to database...")
    conn = psycopg2.connect(DATABASE_URL)
    cursor = conn.cursor()
    
    try:
        cursor.execute("DELETE FROM influencers")
        cursor.execute("DELETE FROM ugc_creators")
        print("Cleared existing data")
        
        print("\nImporting influencers...")
        influencer_count = import_influencers(cursor, wb)
        print(f"Imported {influencer_count} influencers")
        
        print("\nImporting UGC creators...")
        ugc_count = import_ugc_creators(cursor, wb)
        print(f"Imported {ugc_count} UGC creators")
        
        conn.commit()
        print("\nImport completed successfully!")
        
    except Exception as e:
        conn.rollback()
        print(f"Error: {e}")
        raise
    finally:
        cursor.close()
        conn.close()

if __name__ == '__main__':
    main()
